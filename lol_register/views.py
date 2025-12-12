from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.admin.views.decorators import staff_member_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, Q
from django.utils import timezone
from django.contrib import messages
from django.core.paginator import Paginator
from datetime import datetime, timedelta
from decimal import Decimal
import json

from .models import Patient, Visit, VisitProduct, WorkshopOrder, LOLPayment, LOLProductService
from .forms import (
    PatientForm, PatientSearchForm, VisitStep1Form, VisitStep2Form,
    VisitClinicianForm, VisitProductForm, VisitProductUpdateForm,
    WorkshopOrderForm, WorkshopOrderCreateForm, LOLPaymentForm,
    LOLPaymentCreateForm, ExportFilterForm
)


# ============================================================================
# DASHBOARD
# ============================================================================

@staff_member_required
def dashboard(request):
    """Main dashboard with statistics and charts"""
    today = timezone.now().date()
    
    # Calculate date ranges
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)
    
    # Visit statistics
    visits_today = Visit.objects.filter(created_at__date=today).count()
    visits_week = Visit.objects.filter(created_at__date__gte=week_start).count()
    visits_month = Visit.objects.filter(created_at__date__gte=month_start).count()
    
    new_visits = Visit.objects.filter(
        created_at__date__gte=month_start,
        visit_status='New'
    ).count()
    old_visits = Visit.objects.filter(
        created_at__date__gte=month_start,
        visit_status='Old'
    ).count()
    
    # Revenue statistics
    revenue_today = LOLPayment.objects.filter(
        payment_date__date=today
    ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
    
    # Outstanding balance - sum of all visit balances
    all_visits = Visit.objects.all()
    outstanding_balance = sum(v.outstanding_balance for v in all_visits)
    
    # Workshop statistics
    products_to_make = VisitProduct.objects.filter(status='To Make').count()
    products_completed = VisitProduct.objects.filter(
        status__in=['Made/Fitted', 'Repaired']
    ).count()
    
    # Chart data - Gender breakdown
    gender_data = Patient.objects.values('gender').annotate(count=Count('id'))
    gender_labels = [g['gender'] for g in gender_data]
    gender_values = [g['count'] for g in gender_data]
    
    # Chart data - Child/Adult breakdown
    category_data = Patient.objects.values('child_or_adult').annotate(count=Count('id'))
    category_labels = [c['child_or_adult'] for c in category_data]
    category_values = [c['count'] for c in category_data]
    
    # Chart data - Payment method breakdown
    payment_data = LOLPayment.objects.filter(
        payment_date__date__gte=month_start
    ).values('method').annotate(total=Sum('amount_paid'))
    payment_labels = [p['method'] for p in payment_data]
    payment_values = [float(p['total'] or 0) for p in payment_data]
    
    # Chart data - Top 5 services
    top_services = VisitProduct.objects.exclude(status='Cancelled').values(
        'product__name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:5]
    service_labels = [s['product__name'] for s in top_services]
    service_values = [s['count'] for s in top_services]
    
    # Chart data - Last 30-day revenue
    daily_revenue = []
    for i in range(29, -1, -1):
        date = today - timedelta(days=i)
        revenue = LOLPayment.objects.filter(
            payment_date__date=date
        ).aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
        daily_revenue.append({
            'date': date.strftime('%b %d'),
            'revenue': float(revenue)
        })
    
    # Recent visits
    recent_visits = Visit.objects.select_related('patient', 'created_by')[:10]
    
    # Recent payments
    recent_payments = LOLPayment.objects.select_related('visit', 'visit__patient')[:10]
    
    context = {
        'title': 'Register Book - Dashboard',
        'visits_today': visits_today,
        'visits_week': visits_week,
        'visits_month': visits_month,
        'new_visits': new_visits,
        'old_visits': old_visits,
        'revenue_today': float(revenue_today),
        'outstanding_balance': float(outstanding_balance),
        'products_to_make': products_to_make,
        'products_completed': products_completed,
        'gender_labels': json.dumps(gender_labels),
        'gender_values': json.dumps(gender_values),
        'category_labels': json.dumps(category_labels),
        'category_values': json.dumps(category_values),
        'payment_labels': json.dumps(payment_labels),
        'payment_values': json.dumps(payment_values),
        'service_labels': json.dumps(service_labels),
        'service_values': json.dumps(service_values),
        'daily_revenue': json.dumps(daily_revenue),
        'recent_visits': recent_visits,
        'recent_payments': recent_payments,
    }
    
    return render(request, 'lol_register/dashboard.html', context)


# ============================================================================
# PATIENTS
# ============================================================================

@staff_member_required
def patient_list(request):
    """List all patients with search and pagination"""
    patients = Patient.objects.all()
    
    # Search
    search_query = request.GET.get('q', '')
    if search_query:
        patients = patients.filter(
            Q(full_name__icontains=search_query) |
            Q(unique_code__icontains=search_query) |
            Q(contact__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(patients, 25)
    page = request.GET.get('page')
    patients = paginator.get_page(page)
    
    context = {
        'title': 'Patients',
        'patients': patients,
        'search_query': search_query,
    }
    
    return render(request, 'lol_register/patient/patient_list.html', context)


@staff_member_required
def patient_create(request):
    """Create new patient"""
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            patient = form.save()
            messages.success(request, f'Patient {patient.full_name} created successfully! Code: {patient.unique_code}')
            return redirect('lol_register:patient_detail', pk=patient.pk)
    else:
        form = PatientForm()
    
    context = {
        'title': 'Register New Patient',
        'form': form,
    }
    
    return render(request, 'lol_register/patient/patient_create.html', context)


@staff_member_required
def patient_detail(request, pk):
    """View patient details and visit history"""
    patient = get_object_or_404(Patient, pk=pk)
    visits = patient.visits.all()[:10]
    
    context = {
        'title': f'Patient: {patient.full_name}',
        'patient': patient,
        'visits': visits,
    }
    
    return render(request, 'lol_register/patient/patient_detail.html', context)


@staff_member_required
def patient_update(request, pk):
    """Update patient details"""
    patient = get_object_or_404(Patient, pk=pk)
    
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            messages.success(request, f'Patient {patient.full_name} updated successfully!')
            return redirect('lol_register:patient_detail', pk=patient.pk)
    else:
        form = PatientForm(instance=patient)
    
    context = {
        'title': f'Edit Patient: {patient.full_name}',
        'form': form,
        'patient': patient,
    }
    
    return render(request, 'lol_register/patient/patient_update.html', context)


@staff_member_required
def patient_delete(request, pk):
    """Delete patient (only if no visits)"""
    patient = get_object_or_404(Patient, pk=pk)
    
    if request.method == 'POST':
        if patient.visits.exists():
            messages.error(request, 'Cannot delete patient with existing visits.')
            return redirect('lol_register:patient_detail', pk=patient.pk)
        
        patient.delete()
        messages.success(request, 'Patient deleted successfully.')
        return redirect('lol_register:patient_list')
    
    return redirect('lol_register:patient_detail', pk=patient.pk)


# ============================================================================
# VISITS
# ============================================================================

@staff_member_required
def visit_list(request):
    """List all visits with filtering"""
    visits = Visit.objects.select_related('patient', 'created_by').all()
    
    # Filter by date
    date_filter = request.GET.get('date')
    if date_filter:
        visits = visits.filter(created_at__date=date_filter)
    
    # Filter by status
    status_filter = request.GET.get('status')
    if status_filter:
        visits = visits.filter(visit_status=status_filter)
    
    # Search
    search_query = request.GET.get('q', '')
    if search_query:
        visits = visits.filter(
            Q(patient__full_name__icontains=search_query) |
            Q(patient__unique_code__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(visits, 25)
    page = request.GET.get('page')
    visits = paginator.get_page(page)
    
    context = {
        'title': 'Visits',
        'visits': visits,
        'search_query': search_query,
        'date_filter': date_filter or '',
        'status_filter': status_filter or '',
    }
    
    return render(request, 'lol_register/visit/visit_list.html', context)


@staff_member_required
def visit_create_step1(request):
    """Step 1: Select patient for new visit"""
    patients = Patient.objects.all()[:20]
    
    if request.method == 'POST':
        patient_id = request.POST.get('patient_id')
        if patient_id:
            return redirect('lol_register:visit_create_step2', patient_id=patient_id)
        messages.error(request, 'Please select a patient.')
    
    context = {
        'title': 'New Visit - Step 1: Select Patient',
        'patients': patients,
    }
    
    return render(request, 'lol_register/visit/visit_create_step1_patient.html', context)


@staff_member_required
def visit_create_step2(request, patient_id):
    """Step 2: Create visit for selected patient"""
    patient = get_object_or_404(Patient, pk=patient_id)
    
    if request.method == 'POST':
        # Create visit - snapshot is auto-populated in model save()
        visit = Visit.objects.create(
            patient=patient,
            created_by=request.user
        )
        messages.success(request, f'Visit created for {patient.full_name}')
        return redirect('lol_register:visit_detail', pk=visit.pk)
    
    context = {
        'title': f'New Visit - Step 2: Confirm for {patient.full_name}',
        'patient': patient,
    }
    
    return render(request, 'lol_register/visit/visit_create_step2_reception.html', context)


@staff_member_required
def visit_detail(request, pk):
    """View visit details including products and payments"""
    visit = get_object_or_404(Visit.objects.select_related('patient', 'created_by'), pk=pk)
    visit_products = visit.visit_products.select_related('product').all()
    payments = visit.payments.all()
    
    context = {
        'title': f'Visit: {visit.patient.unique_code} - {visit.created_at.strftime("%Y-%m-%d")}',
        'visit': visit,
        'visit_products': visit_products,
        'payments': payments,
        'products': LOLProductService.objects.filter(active=True),
    }
    
    return render(request, 'lol_register/visit/visit_detail.html', context)


@staff_member_required
def visit_update_clinician(request, pk):
    """Clinician updates diagnosis and treatment"""
    visit = get_object_or_404(Visit, pk=pk)
    
    if request.method == 'POST':
        form = VisitClinicianForm(request.POST, instance=visit)
        if form.is_valid():
            form.save()
            messages.success(request, 'Visit updated successfully!')
            return redirect('lol_register:visit_detail', pk=visit.pk)
    else:
        form = VisitClinicianForm(instance=visit)
    
    context = {
        'title': f'Update Visit: {visit.patient.unique_code}',
        'form': form,
        'visit': visit,
    }
    
    return render(request, 'lol_register/visit/visit_update_clinician.html', context)


# ============================================================================
# VISIT PRODUCTS
# ============================================================================

@staff_member_required
def visit_products_list(request, visit_id):
    """List products for a visit"""
    visit = get_object_or_404(Visit, pk=visit_id)
    products = visit.visit_products.select_related('product').all()
    
    context = {
        'title': f'Products for Visit: {visit.patient.unique_code}',
        'visit': visit,
        'products': products,
    }
    
    return render(request, 'lol_register/visit_product/visit_products_list.html', context)


@staff_member_required
def visit_product_add(request, visit_id):
    """Add product to visit"""
    visit = get_object_or_404(Visit, pk=visit_id)
    
    if request.method == 'POST':
        form = VisitProductForm(request.POST)
        if form.is_valid():
            visit_product = form.save(commit=False)
            visit_product.visit = visit
            visit_product.price_at_time = visit_product.product.price
            visit_product.save()
            messages.success(request, f'Added {visit_product.product.name} to visit.')
            return redirect('lol_register:visit_detail', pk=visit.pk)
    else:
        form = VisitProductForm()
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # AJAX request - return JSON
        if request.method == 'POST' and form.is_valid():
            return JsonResponse({'success': True})
        return JsonResponse({'success': False, 'errors': form.errors})
    
    context = {
        'title': f'Add Product to Visit: {visit.patient.unique_code}',
        'form': form,
        'visit': visit,
    }
    
    return render(request, 'lol_register/visit_product/modal_product_add.html', context)


@staff_member_required
def visit_product_update(request, pk):
    """Update visit product status"""
    visit_product = get_object_or_404(VisitProduct, pk=pk)
    
    if request.method == 'POST':
        form = VisitProductUpdateForm(request.POST, instance=visit_product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product status updated.')
            return redirect('lol_register:visit_detail', pk=visit_product.visit.pk)
    else:
        form = VisitProductUpdateForm(instance=visit_product)
    
    context = {
        'title': f'Update Product: {visit_product.product.name}',
        'form': form,
        'visit_product': visit_product,
    }
    
    return render(request, 'lol_register/visit_product/visit_product_update.html', context)


@staff_member_required
def visit_product_delete(request, pk):
    """Delete visit product"""
    visit_product = get_object_or_404(VisitProduct, pk=pk)
    visit_id = visit_product.visit.pk
    
    if request.method == 'POST':
        visit_product.delete()
        messages.success(request, 'Product removed from visit.')
        return redirect('lol_register:visit_detail', pk=visit_id)
    
    return redirect('lol_register:visit_detail', pk=visit_id)


# ============================================================================
# PAYMENTS (CASH BOOK)
# ============================================================================

@staff_member_required
def payment_list(request):
    """List all payments (Cash Book)"""
    payments = LOLPayment.objects.select_related('visit', 'visit__patient').all()
    
    # Date filter
    date_filter = request.GET.get('date')
    if date_filter:
        payments = payments.filter(payment_date__date=date_filter)
    
    # Method filter
    method_filter = request.GET.get('method')
    if method_filter:
        payments = payments.filter(method=method_filter)
    
    # Search
    search_query = request.GET.get('q', '')
    if search_query:
        payments = payments.filter(
            Q(visit__patient__full_name__icontains=search_query) |
            Q(visit__patient__unique_code__icontains=search_query) |
            Q(mpesa_transaction_id__icontains=search_query)
        )
    
    # Calculate totals
    total_expected = payments.aggregate(total=Sum('expected_pay'))['total'] or Decimal('0')
    total_paid = payments.aggregate(total=Sum('amount_paid'))['total'] or Decimal('0')
    total_balance = payments.aggregate(total=Sum('balance'))['total'] or Decimal('0')
    
    # Pagination
    paginator = Paginator(payments, 25)
    page = request.GET.get('page')
    payments = paginator.get_page(page)
    
    context = {
        'title': 'Cash Book (Payments)',
        'payments': payments,
        'search_query': search_query,
        'date_filter': date_filter or '',
        'method_filter': method_filter or '',
        'total_expected': float(total_expected),
        'total_paid': float(total_paid),
        'total_balance': float(total_balance),
    }
    
    return render(request, 'lol_register/payment/payment_list.html', context)


@staff_member_required
def payment_create(request):
    """Create new payment"""
    if request.method == 'POST':
        form = LOLPaymentForm(request.POST)
        if form.is_valid():
            payment = form.save()
            messages.success(request, 'Payment recorded successfully!')
            return redirect('lol_register:payment_list')
    else:
        form = LOLPaymentForm()
    
    context = {
        'title': 'Record Payment',
        'form': form,
        'visits': Visit.objects.select_related('patient').order_by('-created_at')[:50],
    }
    
    return render(request, 'lol_register/payment/payment_create.html', context)


@staff_member_required
def payment_create_for_visit(request, visit_id):
    """Create payment for specific visit"""
    visit = get_object_or_404(Visit, pk=visit_id)
    
    if request.method == 'POST':
        form = LOLPaymentCreateForm(request.POST)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.visit = visit
            payment.save()
            messages.success(request, 'Payment recorded successfully!')
            return redirect('lol_register:visit_detail', pk=visit.pk)
    else:
        # Pre-fill expected pay with outstanding balance
        initial = {'expected_pay': visit.outstanding_balance}
        form = LOLPaymentCreateForm(initial=initial)
    
    context = {
        'title': f'Record Payment for {visit.patient.unique_code}',
        'form': form,
        'visit': visit,
    }
    
    return render(request, 'lol_register/payment/payment_create.html', context)


@staff_member_required
def payment_update(request, pk):
    """Update payment"""
    payment = get_object_or_404(LOLPayment, pk=pk)
    
    if request.method == 'POST':
        form = LOLPaymentForm(request.POST, instance=payment)
        if form.is_valid():
            form.save()
            messages.success(request, 'Payment updated successfully!')
            return redirect('lol_register:payment_list')
    else:
        form = LOLPaymentForm(instance=payment)
    
    context = {
        'title': f'Edit Payment',
        'form': form,
        'payment': payment,
    }
    
    return render(request, 'lol_register/payment/payment_update.html', context)


@staff_member_required
def payment_delete(request, pk):
    """Delete payment"""
    payment = get_object_or_404(LOLPayment, pk=pk)
    
    if request.method == 'POST':
        payment.delete()
        messages.success(request, 'Payment deleted successfully.')
        return redirect('lol_register:payment_list')
    
    return redirect('lol_register:payment_list')


# ============================================================================
# WORKSHOP ORDERS
# ============================================================================

@staff_member_required
def workshop_list(request):
    """List all workshop orders"""
    orders = WorkshopOrder.objects.select_related(
        'visit_product', 'visit_product__visit', 'visit_product__visit__patient',
        'visit_product__product', 'assigned_to'
    ).all()
    
    # Status filter
    status_filter = request.GET.get('status')
    if status_filter:
        orders = orders.filter(status=status_filter)
    
    # Search
    search_query = request.GET.get('q', '')
    if search_query:
        orders = orders.filter(
            Q(visit_product__visit__patient__full_name__icontains=search_query) |
            Q(visit_product__visit__patient__unique_code__icontains=search_query) |
            Q(visit_product__product__name__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(orders, 25)
    page = request.GET.get('page')
    orders = paginator.get_page(page)
    
    # Statistics
    to_make = WorkshopOrder.objects.filter(status='To Make').count()
    in_progress = WorkshopOrder.objects.filter(status='In Progress').count()
    made = WorkshopOrder.objects.filter(status='Made').count()
    fitted = WorkshopOrder.objects.filter(status='Fitted').count()
    
    context = {
        'title': 'Workshop Orders',
        'orders': orders,
        'search_query': search_query,
        'status_filter': status_filter or '',
        'to_make': to_make,
        'in_progress': in_progress,
        'made': made,
        'fitted': fitted,
    }
    
    return render(request, 'lol_register/workshop/workshop_order_list.html', context)


@staff_member_required
def workshop_create(request):
    """Create new workshop order"""
    if request.method == 'POST':
        form = WorkshopOrderCreateForm(request.POST)
        if form.is_valid():
            order = form.save()
            messages.success(request, 'Workshop order created successfully!')
            return redirect('lol_register:workshop_detail', pk=order.pk)
    else:
        form = WorkshopOrderCreateForm()
    
    context = {
        'title': 'Create Workshop Order',
        'form': form,
    }
    
    return render(request, 'lol_register/workshop/workshop_order_create.html', context)


@staff_member_required
def workshop_create_for_product(request, visit_product_id):
    """Create workshop order for specific visit product"""
    visit_product = get_object_or_404(VisitProduct, pk=visit_product_id)
    
    # Check if order already exists
    if hasattr(visit_product, 'workshop_order'):
        messages.warning(request, 'Workshop order already exists for this product.')
        return redirect('lol_register:workshop_detail', pk=visit_product.workshop_order.pk)
    
    if request.method == 'POST':
        form = WorkshopOrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.visit_product = visit_product
            order.save()
            messages.success(request, 'Workshop order created successfully!')
            return redirect('lol_register:workshop_detail', pk=order.pk)
    else:
        form = WorkshopOrderForm()
    
    context = {
        'title': f'Create Workshop Order: {visit_product.product.name}',
        'form': form,
        'visit_product': visit_product,
    }
    
    return render(request, 'lol_register/workshop/workshop_order_create.html', context)


@staff_member_required
def workshop_detail(request, pk):
    """View workshop order details"""
    order = get_object_or_404(WorkshopOrder.objects.select_related(
        'visit_product', 'visit_product__visit', 'visit_product__visit__patient',
        'visit_product__product', 'assigned_to'
    ), pk=pk)
    
    context = {
        'title': f'Workshop Order: {order.visit_product.product.name}',
        'order': order,
    }
    
    return render(request, 'lol_register/workshop/workshop_order_detail.html', context)


@staff_member_required
def workshop_update(request, pk):
    """Update workshop order"""
    order = get_object_or_404(WorkshopOrder, pk=pk)
    
    if request.method == 'POST':
        form = WorkshopOrderForm(request.POST, instance=order)
        if form.is_valid():
            updated_order = form.save()
            
            # Update visit product status based on workshop order status
            if updated_order.status in ['Made', 'Fitted']:
                updated_order.visit_product.status = 'Made/Fitted'
                updated_order.visit_product.save()
            elif updated_order.status == 'Repaired':
                updated_order.visit_product.status = 'Repaired'
                updated_order.visit_product.save()
            
            messages.success(request, 'Workshop order updated successfully!')
            return redirect('lol_register:workshop_detail', pk=order.pk)
    else:
        form = WorkshopOrderForm(instance=order)
    
    context = {
        'title': f'Update Workshop Order: {order.visit_product.product.name}',
        'form': form,
        'order': order,
    }
    
    return render(request, 'lol_register/workshop/workshop_order_update.html', context)


# ============================================================================
# EXPORTS
# ============================================================================

@staff_member_required
def export_visits(request):
    """Export visits to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    # Get date range
    date_range = request.GET.get('range', 'month')
    today = timezone.now().date()
    
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    else:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            start_date = today.replace(day=1)
            end_date = today
    
    visits = Visit.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).select_related('patient', 'created_by')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Visits"
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="27baf1", end_color="27baf1", fill_type="solid")
    
    # Headers
    headers = ['Date', 'Patient Code', 'Patient Name', 'Gender', 'Age', 'Category', 
               'Status', 'Diagnosis', 'Treatment', 'Next Visit', 'Created By']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, visit in enumerate(visits, 2):
        ws.cell(row=row, column=1, value=visit.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=2, value=visit.patient.unique_code)
        ws.cell(row=row, column=3, value=visit.snapshot_full_name)
        ws.cell(row=row, column=4, value=visit.snapshot_gender)
        ws.cell(row=row, column=5, value=f"{visit.snapshot_age_years}y {visit.snapshot_age_months}m")
        ws.cell(row=row, column=6, value=visit.snapshot_child_or_adult)
        ws.cell(row=row, column=7, value=visit.visit_status)
        ws.cell(row=row, column=8, value=visit.diagnosis or '')
        ws.cell(row=row, column=9, value=visit.treatment_notes or '')
        ws.cell(row=row, column=10, value=visit.next_visit.strftime('%Y-%m-%d') if visit.next_visit else '')
        ws.cell(row=row, column=11, value=visit.created_by.get_full_name() if visit.created_by else '')
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="visits_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    
    return response


@staff_member_required
def export_payments(request):
    """Export payments to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Get date range
    date_range = request.GET.get('range', 'month')
    today = timezone.now().date()
    
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    else:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            start_date = today.replace(day=1)
            end_date = today
    
    payments = LOLPayment.objects.filter(
        payment_date__date__gte=start_date,
        payment_date__date__lte=end_date
    ).select_related('visit', 'visit__patient')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="16134a", end_color="16134a", fill_type="solid")
    
    # Headers
    headers = ['Date', 'Patient Code', 'Patient Name', 'Method', 'Expected', 
               'Paid', 'Balance', 'M-Pesa ID', 'Paid By']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, payment in enumerate(payments, 2):
        ws.cell(row=row, column=1, value=payment.payment_date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=2, value=payment.visit.patient.unique_code)
        ws.cell(row=row, column=3, value=payment.visit.patient.full_name)
        ws.cell(row=row, column=4, value=payment.method)
        ws.cell(row=row, column=5, value=float(payment.expected_pay))
        ws.cell(row=row, column=6, value=float(payment.amount_paid or 0))
        ws.cell(row=row, column=7, value=float(payment.balance))
        ws.cell(row=row, column=8, value=payment.mpesa_transaction_id or '')
        ws.cell(row=row, column=9, value=payment.paid_by or '')
    
    # Totals row
    last_row = len(payments) + 2
    ws.cell(row=last_row, column=4, value='TOTALS:').font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=float(sum(p.expected_pay for p in payments))).font = Font(bold=True)
    ws.cell(row=last_row, column=6, value=float(sum(p.amount_paid or 0 for p in payments))).font = Font(bold=True)
    ws.cell(row=last_row, column=7, value=float(sum(p.balance for p in payments))).font = Font(bold=True)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="payments_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    
    return response


@staff_member_required
def export_products(request):
    """Export products/manufacturing to Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    
    # Get date range
    date_range = request.GET.get('range', 'month')
    today = timezone.now().date()
    
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    else:
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        if start_date and end_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            start_date = today.replace(day=1)
            end_date = today
    
    visit_products = VisitProduct.objects.filter(
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).select_related('visit', 'visit__patient', 'product')
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="27baf1", end_color="27baf1", fill_type="solid")
    
    # Headers
    headers = ['Date', 'Patient Code', 'Patient Name', 'Product', 'Quantity', 
               'Unit Price', 'Subtotal', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Data
    for row, vp in enumerate(visit_products, 2):
        ws.cell(row=row, column=1, value=vp.created_at.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=2, value=vp.visit.patient.unique_code)
        ws.cell(row=row, column=3, value=vp.visit.patient.full_name)
        ws.cell(row=row, column=4, value=vp.product.name)
        ws.cell(row=row, column=5, value=vp.quantity)
        ws.cell(row=row, column=6, value=float(vp.price_at_time))
        ws.cell(row=row, column=7, value=float(vp.subtotal))
        ws.cell(row=row, column=8, value=vp.status)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="products_{start_date}_{end_date}.xlsx"'
    wb.save(response)
    
    return response


# ============================================================================
# SEARCH
# ============================================================================

@staff_member_required
def search(request):
    """Global search across patients and visits"""
    query = request.GET.get('q', '')
    
    patients = []
    visits = []
    
    if query:
        patients = Patient.objects.filter(
            Q(full_name__icontains=query) |
            Q(unique_code__icontains=query) |
            Q(contact__icontains=query)
        )[:20]
        
        visits = Visit.objects.filter(
            Q(patient__full_name__icontains=query) |
            Q(patient__unique_code__icontains=query)
        ).select_related('patient')[:20]
    
    context = {
        'title': 'Search Results',
        'query': query,
        'patients': patients,
        'visits': visits,
    }
    
    return render(request, 'lol_register/search.html', context)


# ============================================================================
# API ENDPOINTS
# ============================================================================

@staff_member_required
def api_patient_search(request):
    """API endpoint for patient search (AJAX)"""
    query = request.GET.get('q', '')
    
    if len(query) < 2:
        return JsonResponse({'results': []})
    
    patients = Patient.objects.filter(
        Q(full_name__icontains=query) |
        Q(unique_code__icontains=query)
    )[:10]
    
    results = [{
        'id': p.id,
        'text': f"{p.full_name} ({p.unique_code})",
        'full_name': p.full_name,
        'unique_code': p.unique_code,
        'gender': p.gender,
        'age': f"{p.age_years}y {p.age_months}m",
    } for p in patients]
    
    return JsonResponse({'results': results})


@staff_member_required
def api_product_price(request, product_id):
    """API endpoint to get product price"""
    try:
        product = LOLProductService.objects.get(pk=product_id, active=True)
        return JsonResponse({
            'success': True,
            'price': float(product.price),
            'name': product.name
        })
    except LOLProductService.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Product not found'})


@staff_member_required
def api_dashboard_stats(request):
    """API endpoint for real-time dashboard stats"""
    today = timezone.now().date()
    
    stats = {
        'visits_today': Visit.objects.filter(created_at__date=today).count(),
        'revenue_today': float(
            LOLPayment.objects.filter(payment_date__date=today)
            .aggregate(total=Sum('amount_paid'))['total'] or 0
        ),
        'products_to_make': VisitProduct.objects.filter(status='To Make').count(),
    }
    
    return JsonResponse(stats)
